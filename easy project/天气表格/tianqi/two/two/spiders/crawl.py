# -*- coding: utf-8 -*-
import scrapy
import json
from openpyxl import Workbook


class CrawlSpider(scrapy.Spider):
    name = 'crawl'
    allowed_domains = ['xxfb.mwr.cn']
    start_urls = ['http://xxfb.mwr.cn/hydroSearch/greatRiver',
                  'http://xxfb.mwr.cn/hydroSearch/greatRsvr',
                  'http://xxfb.mwr.cn/hydroSearch/pointHydroInfo']

    def parse(self, response):
        url = response.url
        a = response.text
        data = json.loads(a)["result"]["data"]
        if 'greatRiver' in url:
            return self.get_greatRiver(data)
        if 'greatRsvr' in url:
            return self.get_greatRsvr(data)
        else:
            return self.get_pointHydroInfo(data)

    def get_greatRiver(self, data):

        data = [[i["poiBsnm"].strip(),
                 i["poiAddv"].strip(),
                 i["rvnm"].strip(),
                 i["stnm"].strip(),
                 i["dateTime"],
                 i["zl"],
                 i["ql"],
                 i["wrz"]] for i in data]
        wb = Workbook()
        ws = wb.active
        ws.append(['流域', '行政区', '河名',
                   '站名', '时间', '水位(米)',
                   '流量(米3/秒)', '警戒水位(米)'])
        for i in data:
            ws.append(i)
        wb.save('greatRiver.xlsx')

    def get_greatRsvr(self, data):

        data = [[i["poiBsnm"].strip(),
                 i["poiAddv"].strip(),
                 i["rvnm"].strip(),
                 i["stnm"].strip(),
                 i["rz"],
                 i["wl"],
                 i["inq"],
                 i["damel"]] for i in data]
        wb = Workbook()
        ws = wb.active
        ws.append(['流域', '行政区', '河名',
                   '库名', '库水位(米)',
                   '蓄水量(百万3)', '入库(米3/秒)', '坝顶高程(米)'
                   ])
        for i in data:
            ws.append(i)
        wb.save('greatRsvr.xlsx')

    def get_pointHydroInfo(self, data):

        data = [[i["poiBsnm"].strip(),
                 i["poiAddv"].strip(),
                 i["rvnm"].strip(),
                 i["stnm"].strip(),
                 i["dateTime"],
                 i["dyp"],
                 i["wth"], ] for i in data]
        wb = Workbook()
        ws = wb.active
        ws.append(['流域', '行政区', '河名', '站名', '日期', '日雨量(毫米)', '天气'])
        for i in data:
            ws.append(i)
        wb.save('pointHydroInfo.xlsx')
